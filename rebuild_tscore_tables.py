"""Rebuild PAI T-score tables from ground truth Excel + reference derivation.

Full/standalone scales: use ground truth Excel data (Table A.1) directly,
with OLS-extrapolated values for raw scores beyond the table.

Subscales: derive M/SD from two validated reference clients (Eve and Greg).
For subscales where both clients have identical raw scores, use Table A.2 fallback.
"""
import json
import os

# ── Load ground truth from Excel ──
try:
    import openpyxl
except ImportError:
    raise SystemExit("pip install openpyxl")

GT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       'PAI docs', 'PAI_RAW-to-T-score_GROUND TRUTH.xlsx')
wb = openpyxl.load_workbook(GT_PATH, data_only=True)
ws = wb['Sheet1']

# Parse headers
headers = [str(c.value).strip() if c.value else ''
           for c in next(ws.iter_rows(min_row=1, max_row=1))]

# Parse data (row 4+ = raw 0+)
GT_DATA = {}  # scale -> {raw: T}
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    raw = row[0]
    if raw is None:
        continue
    raw = int(raw)
    for i, val in enumerate(row[1:], 1):
        if val is not None and i < len(headers):
            scale = headers[i]
            if scale not in GT_DATA:
                GT_DATA[scale] = {}
            GT_DATA[scale][raw] = int(val)

# Fix known data entry errors
if 'INF' in GT_DATA and GT_DATA['INF'].get(17) == 10:
    GT_DATA['INF'][17] = 106
# ANX column is off by +1 (GT says T=34+raw, PAR reports confirm T=33+raw)
if 'ANX' in GT_DATA:
    GT_DATA['ANX'] = {r: t - 1 for r, t in GT_DATA['ANX'].items()}

print(f"Loaded ground truth: {len(GT_DATA)} scales from Excel")

# ── Reference data from PAR scoring reports ──
EVE_REF = {
    'ICN': (9, 61), 'INF': (5, 59), 'NIM': (1, 47), 'PIM': (13, 45),
    'SOM': (12, 51), 'ANX': (19, 52), 'ARD': (21, 51), 'DEP': (10, 45),
    'MAN': (35, 63), 'PAR': (22, 54), 'SCZ': (14, 50), 'BOR': (29, 61),
    'ANT': (17, 54), 'ALC': (6, 52), 'DRG': (5, 52), 'AGG': (5, 38),
    'SUI': (4, 51), 'STR': (5, 48), 'NON': (4, 48), 'RXR': (14, 51),
    'DOM': (18, 45), 'WRM': (17, 38),
    'SOM-C': (1, 46), 'SOM-S': (4, 49), 'SOM-H': (7, 57),
    'ANX-C': (7, 52), 'ANX-A': (6, 49), 'ANX-P': (6, 55),
    'ARD-O': (12, 57), 'ARD-P': (5, 45), 'ARD-T': (4, 50),
    'DEP-C': (1, 40), 'DEP-A': (2, 44), 'DEP-P': (7, 53),
    'MAN-A': (9, 57), 'MAN-G': (11, 56), 'MAN-I': (15, 67),
    'PAR-H': (10, 57), 'PAR-P': (2, 45), 'PAR-R': (10, 58),
    'SCZ-P': (3, 46), 'SCZ-S': (6, 51), 'SCZ-T': (5, 52),
    'BOR-A': (9, 63), 'BOR-I': (8, 59), 'BOR-N': (8, 59), 'BOR-S': (4, 53),
    'ANT-A': (7, 55), 'ANT-E': (3, 49), 'ANT-S': (7, 56),
    'AGG-A': (4, 45), 'AGG-V': (1, 34), 'AGG-P': (0, 42),
}

GREG_REF = {
    'ICN': (8, 58), 'INF': (6, 63), 'NIM': (10, 81), 'PIM': (14, 48),
    'SOM': (19, 58), 'ANX': (24, 57), 'ARD': (19, 49), 'DEP': (54, 92),
    'MAN': (21, 48), 'PAR': (62, 100), 'SCZ': (25, 64), 'BOR': (43, 75),
    'ANT': (53, 94), 'ALC': (3, 47), 'DRG': (0, 42), 'AGG': (53, 95),
    'SUI': (12, 68), 'STR': (15, 71), 'NON': (21, 94), 'RXR': (18, 59),
    'DOM': (21, 51), 'WRM': (0, 30),
    'SOM-C': (0, 43), 'SOM-S': (12, 70), 'SOM-H': (7, 57),
    'ANX-C': (8, 55), 'ANX-A': (9, 57), 'ANX-P': (7, 58),
    'ARD-O': (6, 41), 'ARD-P': (7, 51), 'ARD-T': (6, 55),
    'DEP-C': (19, 93), 'DEP-A': (21, 96), 'DEP-P': (14, 69),
    'MAN-A': (2, 35), 'MAN-G': (4, 40), 'MAN-I': (15, 67),
    'PAR-H': (23, 95), 'PAR-P': (20, 98), 'PAR-R': (19, 83),
    'SCZ-P': (2, 43), 'SCZ-S': (23, 94), 'SCZ-T': (0, 37),
    'BOR-A': (13, 75), 'BOR-I': (9, 62), 'BOR-N': (16, 84), 'BOR-S': (5, 57),
    'ANT-A': (24, 93), 'ANT-E': (17, 95), 'ANT-S': (12, 70),
    'AGG-A': (18, 84), 'AGG-V': (18, 82), 'AGG-P': (17, 100),
}

# ── Fallback M/SD for subscales where Eve and Greg have identical raw scores ──
FALLBACK_STATS = {
    'SOM-H': (4.09, 4.25),
    'MAN-I': (7.92, 4.27),
}

# ── Extended ground truth (raw 40+ from manual Table A.1) ──
GT_EXTENDED = {
    'SOM': [79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101,102,103,104,105,106,107,108,109,109,110],
    'ANX': [72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101,102,103],
    'ARD': [74,75,77,78,79,80,81,83,84,85,86,87,89,90,91,92,93,95,96,97,98,100,101,102,103,104,106,107,108,109,110,112,113],
    'DEP': [77,78,79,80,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,100,101,102,103,104,105,106,107,108,109,110,111],
    'MAN': [68,70,71,72,73,74,75,76,77,78,79,80,81,83,84,85,86,87,88,89,90,91,92,93,94,96,97,98,99,100,101,102,103],
    'PAR': [75,76,77,78,79,81,82,83,84,85,86,87,89,90,91,92,93,94,96,97,98,99,100,101,102,104,105,106,107,108,109,110,112],
    'SCZ': [83,85,86,87,89,90,91,92,94,95,96,98,99,100,101,103,104,105,106,108,109,110,112,113,114,115,117,118,119,121,122,123,124],
    'BOR': [72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101,102,103,104],
    'ANT': [79,81,82,83,84,85,86,87,88,89,90,92,93,94,95,96,97,98,99,100,101,103,104,105,106,107,108,109,110,111,112,113,115],
    'AGG': [80,81,82,83,85,86,87,88,89,91,92,93,94,95,97],
}

# Merge extended GT into GT_DATA
for scale, ext_values in GT_EXTENDED.items():
    if scale in GT_DATA:
        start_raw = max(GT_DATA[scale].keys()) + 1
        for i, t in enumerate(ext_values):
            GT_DATA[scale][start_raw + i] = t

# ── Max raw scores ──
MAX_RAW = {
    'SOM': 72, 'ANX': 72, 'ARD': 72, 'DEP': 72, 'MAN': 72,
    'PAR': 72, 'SCZ': 72, 'BOR': 72, 'ANT': 72,
    'ALC': 36, 'DRG': 36, 'AGG': 54,
    'SUI': 36, 'STR': 24, 'NON': 24, 'RXR': 24,
    'DOM': 36, 'WRM': 36,
    'INF': 24, 'NIM': 27, 'PIM': 27,
    'SOM-C': 24, 'SOM-S': 24, 'SOM-H': 24,
    'ANX-C': 24, 'ANX-A': 24, 'ANX-P': 24,
    'ARD-O': 24, 'ARD-P': 24, 'ARD-T': 24,
    'DEP-C': 24, 'DEP-A': 24, 'DEP-P': 24,
    'MAN-A': 24, 'MAN-G': 24, 'MAN-I': 24,
    'PAR-H': 24, 'PAR-P': 24, 'PAR-R': 24,
    'SCZ-P': 24, 'SCZ-S': 24, 'SCZ-T': 24,
    'ANT-A': 24, 'ANT-E': 24, 'ANT-S': 24,
    'BOR-A': 18, 'BOR-I': 18, 'BOR-N': 18, 'BOR-S': 18,
    'AGG-A': 18, 'AGG-V': 18, 'AGG-P': 18,
}


def ols_fit(points):
    """Fit y = a + b*x via OLS. Returns (a, b)."""
    n = len(points)
    sx = sum(x for x, y in points)
    sy = sum(y for x, y in points)
    sxx = sum(x * x for x, y in points)
    sxy = sum(x * y for x, y in points)
    b = (n * sxy - sx * sy) / (n * sxx - sx * sx)
    a = (sy - b * sx) / n
    return a, b


def derive_from_refs(scale):
    """Derive intercept and slope from two reference points."""
    eve_raw, eve_t = EVE_REF[scale]
    greg_raw, greg_t = GREG_REF[scale]

    if eve_raw == greg_raw:
        if scale in FALLBACK_STATS:
            M, SD = FALLBACK_STATS[scale]
            return 50 - 10 * M / SD, 10 / SD, "fallback"
        return None, None, None

    slope = (greg_t - eve_t) / (greg_raw - eve_raw)
    if abs(slope) < 0.001:
        return None, None, None

    intercept = eve_t - slope * eve_raw
    return intercept, slope, "derived"


# ── Generate tables ──
print("\nGENERATING T-SCORE TABLES")
print("=" * 70)

tables = {}
all_scales = sorted(set(list(EVE_REF.keys())) - {'ICN'})

for scale in all_scales:
    max_raw = MAX_RAW.get(scale, 24)
    table = {}

    if scale in GT_DATA:
        # ── Full/standalone scale: use ground truth + OLS extrapolation ──
        gt = GT_DATA[scale]
        gt_max = max(gt.keys())

        # Use GT values directly for available raw scores
        for raw in sorted(gt.keys()):
            table[str(raw)] = gt[raw]

        # If we need values beyond GT, extrapolate with OLS
        if max_raw > gt_max:
            points = [(r, t) for r, t in gt.items()]
            a, b = ols_fit(points)
            for raw in range(gt_max + 1, max_raw + 1):
                table[str(raw)] = round(a + b * raw)
            source = f"GT(0-{gt_max})+OLS({gt_max+1}-{max_raw})"
        else:
            source = f"GT(0-{gt_max})"

        # Validate against reference
        eve_raw, eve_t = EVE_REF[scale]
        greg_raw, greg_t = GREG_REF[scale]
        eve_comp = table.get(str(eve_raw))
        greg_comp = table.get(str(greg_raw))
        eve_diff = (eve_comp - eve_t) if eve_comp is not None else None
        greg_diff = (greg_comp - greg_t) if greg_comp is not None else None
        parts = []
        if eve_diff and eve_diff != 0:
            parts.append(f"Eve:{eve_diff:+d}")
        if greg_diff and greg_diff != 0:
            parts.append(f"Greg:{greg_diff:+d}")
        status = ' '.join(parts) if parts else "OK"

    else:
        # ── Subscale: derive from reference points ──
        intercept, slope, src = derive_from_refs(scale)
        if intercept is None:
            print(f"  {scale}: SKIP (can't derive)")
            continue

        for raw in range(max_raw + 1):
            table[str(raw)] = round(intercept + slope * raw)

        source = src
        eve_raw, eve_t = EVE_REF[scale]
        greg_raw, greg_t = GREG_REF[scale]
        eve_comp = table.get(str(eve_raw))
        greg_comp = table.get(str(greg_raw))
        eve_diff = (eve_comp - eve_t) if eve_comp is not None else None
        greg_diff = (greg_comp - greg_t) if greg_comp is not None else None
        parts = []
        if eve_diff and eve_diff != 0:
            parts.append(f"Eve:{eve_diff:+d}")
        if greg_diff and greg_diff != 0:
            parts.append(f"Greg:{greg_diff:+d}")
        status = ' '.join(parts) if parts else "OK"

    print(f"  {scale:<8} ({source})  {status}")

    tables[scale] = {
        'scale_name': scale,
        'raw_to_t': table,
    }

# ── Save ──
output = {'scales': tables}
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           'templates', 'pai', 'tscore_tables.json')
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(output, f, indent=2, ensure_ascii=False)
print(f"\nSaved {len(tables)} scales to {output_path}")

# ── Full validation ──
print(f"\n{'=' * 70}")
print("VALIDATION AGAINST REFERENCE CLIENTS")
print("=" * 70)

import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from validate_pai import EVE_RESPONSES, GREG_RESPONSES

mapping_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            'templates', 'pai', 'mapping.json')
with open(mapping_path, encoding='utf-8') as f:
    mapping = json.load(f)['scales']


def score_scale(scale_abbr, responses):
    if scale_abbr not in mapping:
        return None
    raw = 0
    for entry in mapping[scale_abbr]['items']:
        resp = responses.get(entry['item'])
        if resp is None or entry['keyed'] == 'paired':
            continue
        if entry['keyed'] == 'False':
            raw += (3 - resp)
        else:
            raw += resp
    return raw


def lookup_t(scale_abbr, raw):
    if scale_abbr not in tables:
        return None
    return tables[scale_abbr]['raw_to_t'].get(str(raw))


perfect = raw_m = t_m = total = 0

for name, responses, reference in [
    ("Eve", EVE_RESPONSES, EVE_REF),
    ("Greg", GREG_RESPONSES, GREG_REF),
]:
    print(f"\n  {name} — mismatches only:")
    any_mismatch = False
    for scale in sorted(reference.keys()):
        if scale == 'ICN':
            continue
        ref_raw, ref_t = reference[scale]
        our_raw = score_scale(scale, responses)
        if our_raw is None:
            total += 1
            continue
        our_t = lookup_t(scale, our_raw)
        rok = our_raw == ref_raw
        tok = our_t == ref_t if our_t is not None else False
        if rok: raw_m += 1
        if tok: t_m += 1
        if rok and tok: perfect += 1
        total += 1
        if not (rok and tok):
            any_mismatch = True
            print(f"    {scale:<8} raw={our_raw}/{ref_raw} T={our_t}/{ref_t}")
    if not any_mismatch:
        print(f"    (none)")

print(f"\n{'=' * 70}")
print(f"OVERALL: {perfect}/{total} perfect | Raw: {raw_m}/{total} | T: {t_m}/{total}")
